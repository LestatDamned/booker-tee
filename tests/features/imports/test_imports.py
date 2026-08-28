import asyncio
import traceback
from datetime import date
from decimal import Decimal
from hashlib import sha256
from io import BytesIO
from pathlib import Path
from stat import S_IMODE
from types import SimpleNamespace
from typing import Any, cast
from unittest.mock import AsyncMock
from uuid import UUID, uuid4
from zipfile import ZIP_DEFLATED, ZipFile

import pytest
from fastapi import UploadFile
from openpyxl import Workbook

from app.features.accounts.models import AccountType
from app.features.import_review.domain.queue import REVIEW_QUEUE_STATUSES
from app.features.import_review.repository import ImportReviewRepository
from app.features.imports.documents.attempts import record_failed_parse_attempt
from app.features.imports.documents.commands import upload as upload_module
from app.features.imports.documents.commands.upload import (
    StatementUploadUseCase,
    should_retain_source_file,
    validate_statement_upload,
)
from app.features.imports.documents.errors import (
    UploadIdempotencyConflictError,
    UploadProcessingError,
    UploadTooLargeError,
    UploadValidationError,
)
from app.features.imports.documents.storage import (
    StoredUpload,
    UploadStorage,
    sanitize_upload_filename,
)
from app.features.imports.documents.types import ParseAttemptStatus, UploadedDocumentStatus
from app.features.imports.models import ParseAttempt, RawTransaction, UploadedDocument
from app.features.imports.parsers.extractors import pdf as pdf_extractor_module
from app.features.imports.parsers.extractors.dto import (
    ExtractedStatement,
    ExtractedStatementPageTables,
)
from app.features.imports.parsers.extractors.limits import (
    StatementExtractionLimits,
    StatementResourceLimitError,
)
from app.features.imports.parsers.extractors.pdf import (
    PdfPlumberStatementExtractor,
)
from app.features.imports.parsers.extractors.resolver import (
    StatementExtractorResolver,
)
from app.features.imports.parsers.extractors.xlsx import (
    OpenPyxlStatementExtractor,
)
from app.features.imports.parsers.support.normalization import (
    parse_bank_date,
)
from app.features.imports.statements.deduplication import possible_duplicate_fingerprint
from app.features.imports.statements.dto import RawTransactionDraft
from app.features.imports.statements.process import StatementParseCompletionService
from app.features.imports.statements.raw_transactions import RawTransactionMapper
from app.features.imports.statements.types import RawTransactionStatus
from app.features.ledger.application.ledger_reference_resolver import LedgerReferenceResolver
from app.features.ledger.errors import LedgerPostingError


def test_sanitize_upload_filename_removes_paths_and_unsafe_characters() -> None:
    assert sanitize_upload_filename("../bank statement июнь.pdf") == "bank_statement_.pdf"
    assert sanitize_upload_filename("statement") == "statement"
    assert sanitize_upload_filename("../bank statement июнь.xlsx") == "bank_statement_.xlsx"


async def test_upload_storage_preserves_pdf_bytes(tmp_path: Path) -> None:
    content = b"%PDF-1.4 local fixture bytes"
    upload = UploadFile(file=BytesIO(content), filename="../private-bank-statement.pdf")
    workspace_id = uuid4()
    document_id = uuid4()

    stored = await UploadStorage(tmp_path).save_upload(
        upload,
        workspace_id=workspace_id,
        document_id=document_id,
    )

    assert stored.file_size_bytes == len(content)
    assert stored.sha256_hash == sha256(content).hexdigest()
    assert stored.path.read_bytes() == content
    assert stored.storage_key == f"{workspace_id}/{document_id}/source.pdf"
    assert "private-bank-statement" not in stored.storage_key
    assert S_IMODE(stored.path.stat().st_mode) == 0o600
    assert S_IMODE(tmp_path.stat().st_mode) == 0o700  # noqa: ASYNC240
    assert S_IMODE(stored.path.parent.stat().st_mode) == 0o700
    assert S_IMODE(stored.path.parent.parent.stat().st_mode) == 0o700


async def test_upload_storage_preserves_xlsx_extension(tmp_path: Path) -> None:
    workbook_file = BytesIO()
    workbook = Workbook()
    workbook.active["A1"] = "Дата"
    workbook.save(workbook_file)
    workbook.close()
    content = workbook_file.getvalue()
    upload = UploadFile(file=BytesIO(content), filename="../statement.xlsx")
    workspace_id = uuid4()
    document_id = uuid4()

    stored = await UploadStorage(tmp_path).save_upload(
        upload,
        workspace_id=workspace_id,
        document_id=document_id,
    )

    assert stored.file_size_bytes == len(content)
    assert stored.sha256_hash == sha256(content).hexdigest()
    assert stored.path.read_bytes() == content
    assert stored.storage_key == f"{workspace_id}/{document_id}/source.xlsx"


async def test_upload_storage_rejects_oversized_file_without_leaving_partial(
    tmp_path: Path,
) -> None:
    upload = UploadFile(file=BytesIO(b"too large"), filename="statement.pdf")

    with pytest.raises(UploadTooLargeError):
        await UploadStorage(tmp_path).save_upload(
            upload,
            workspace_id=uuid4(),
            document_id=uuid4(),
            max_bytes=4,
        )

    assert list(tmp_path.rglob("source.pdf")) == []  # noqa: ASYNC240


@pytest.mark.parametrize(
    ("filename", "content", "message"),
    [
        ("statement.pdf", b"not a PDF", "PDF"),
        ("statement.xlsx", b"not an XLSX", "XLSX"),
    ],
)
async def test_upload_storage_rejects_invalid_file_signature(
    tmp_path: Path,
    filename: str,
    content: bytes,
    message: str,
) -> None:
    with pytest.raises(UploadValidationError, match=message):
        await UploadStorage(tmp_path).save_upload(
            UploadFile(file=BytesIO(content), filename=filename),
            workspace_id=uuid4(),
            document_id=uuid4(),
        )

    assert list(tmp_path.rglob("source.*")) == []  # noqa: ASYNC240


async def test_upload_storage_rejects_zip_without_xlsx_structure(tmp_path: Path) -> None:
    content = BytesIO()
    with ZipFile(content, "w", compression=ZIP_DEFLATED) as archive:
        archive.writestr("unrelated.txt", "not a workbook")

    with pytest.raises(UploadValidationError, match="XLSX"):
        await UploadStorage(tmp_path).save_upload(
            UploadFile(file=BytesIO(content.getvalue()), filename="statement.xlsx"),
            workspace_id=uuid4(),
            document_id=uuid4(),
        )

    assert list(tmp_path.rglob("source.xlsx")) == []  # noqa: ASYNC240


async def test_upload_storage_does_not_overwrite_existing_file(tmp_path: Path) -> None:
    storage = UploadStorage(tmp_path)
    workspace_id = uuid4()
    document_id = uuid4()
    original = b"%PDF-1.4 original"
    stored = await storage.save_upload(
        UploadFile(file=BytesIO(original), filename="first.pdf"),
        workspace_id=workspace_id,
        document_id=document_id,
    )

    with pytest.raises(UploadValidationError, match="уже существует"):
        await storage.save_upload(
            UploadFile(file=BytesIO(b"%PDF-1.4 replacement"), filename="second.pdf"),
            workspace_id=workspace_id,
            document_id=document_id,
        )

    assert stored.path.read_bytes() == original


async def test_upload_storage_rejects_path_through_symlink(tmp_path: Path) -> None:
    root_dir = tmp_path / "uploads"
    outside_dir = tmp_path / "outside"
    root_dir.mkdir()
    outside_dir.mkdir()
    workspace_id = uuid4()
    (root_dir / str(workspace_id)).symlink_to(outside_dir, target_is_directory=True)

    with pytest.raises(UploadValidationError, match="storage path"):
        await UploadStorage(root_dir).save_upload(
            UploadFile(file=BytesIO(b"%PDF-1.4"), filename="statement.pdf"),
            workspace_id=workspace_id,
            document_id=uuid4(),
        )

    assert list(outside_dir.iterdir()) == []


async def test_statement_upload_replays_same_idempotent_payload(tmp_path: Path) -> None:
    workspace_id = uuid4()
    account = SimpleNamespace(id=uuid4(), currency="RUB")
    content = b"%PDF-1.4 replay"
    existing = SimpleNamespace(
        id=uuid4(),
        account_id=account.id,
        original_filename="statement.pdf",
        status=UploadedDocumentStatus.REQUIRES_REVIEW,
        sha256_hash=sha256(content).hexdigest(),
    )

    class Accounts:
        async def get_import_account(self, requested_workspace_id, account_id):
            assert requested_workspace_id == workspace_id
            assert account_id == account.id
            return account

    class Imports:
        async def get_document_for_workspace(self, requested_workspace_id, _document_id):
            assert requested_workspace_id == workspace_id
            return existing

    use_case = object.__new__(StatementUploadUseCase)
    use_case.settings = SimpleNamespace(statement_upload_max_bytes=1024)
    use_case.accounts = Accounts()
    use_case.documents = Imports()
    use_case.storage = UploadStorage(tmp_path)

    result = await use_case.upload_statement(
        context=cast(
            Any,
            SimpleNamespace(
                workspace=SimpleNamespace(id=workspace_id),
                user=SimpleNamespace(id=uuid4()),
            ),
        ),
        upload_file=UploadFile(file=BytesIO(content), filename="statement.pdf"),
        account_id=account.id,
        idempotency_key=uuid4(),
    )

    assert result.replayed is True
    assert result.document_id == existing.id
    assert result.document_status is existing.status
    assert result.filename == existing.original_filename
    assert list(tmp_path.rglob("source.pdf")) == []  # noqa: ASYNC240


async def test_statement_upload_rejects_changed_idempotent_payload(tmp_path: Path) -> None:
    workspace_id = uuid4()
    account = SimpleNamespace(id=uuid4(), currency="RUB")
    existing = SimpleNamespace(
        account_id=account.id,
        original_filename="statement.pdf",
        sha256_hash=sha256(b"first").hexdigest(),
    )

    class Accounts:
        async def get_import_account(self, *_args):
            return account

    class Imports:
        async def get_document_for_workspace(self, *_args):
            return existing

    use_case = object.__new__(StatementUploadUseCase)
    use_case.settings = SimpleNamespace(statement_upload_max_bytes=1024)
    use_case.accounts = Accounts()
    use_case.documents = Imports()
    use_case.storage = UploadStorage(tmp_path)

    with pytest.raises(UploadIdempotencyConflictError):
        await use_case.upload_statement(
            context=cast(
                Any,
                SimpleNamespace(
                    workspace=SimpleNamespace(id=workspace_id),
                    user=SimpleNamespace(id=uuid4()),
                ),
            ),
            upload_file=UploadFile(file=BytesIO(b"second"), filename="statement.pdf"),
            account_id=account.id,
            idempotency_key=uuid4(),
        )


@pytest.mark.parametrize("activity_fails", [False, True])
async def test_statement_upload_cleans_stored_file_when_first_transaction_fails(
    tmp_path: Path,
    activity_fails: bool,
) -> None:
    workspace_id = uuid4()
    account = SimpleNamespace(id=uuid4(), currency="RUB")

    class Session:
        def __init__(self) -> None:
            self.commits = 0
            self.rollbacks = 0

        async def commit(self) -> None:
            self.commits += 1

        async def rollback(self) -> None:
            self.rollbacks += 1

    class Accounts:
        async def get_import_account(self, *_args: object) -> object:
            return account

    class Documents:
        async def create_uploaded_document(self, document: object) -> object:
            if not activity_fails:
                raise RuntimeError("document write failed")
            return document

    class UploadStub:
        filename = "../statement.pdf"
        content_type = "application/pdf"

        def __init__(self) -> None:
            self.file = BytesIO(b"%PDF-1.4")

        async def read(self, size: int = -1) -> bytes:
            return self.file.read(size)

        async def seek(self, offset: int) -> None:
            self.file.seek(offset)

    session = Session()
    use_case = object.__new__(StatementUploadUseCase)
    use_case.session = cast(Any, session)
    use_case.settings = cast(
        Any,
        SimpleNamespace(
            statement_upload_max_bytes=1024,
            upload_storage_dir=tmp_path,
        ),
    )
    use_case.accounts = cast(Any, Accounts())
    use_case.documents = cast(Any, Documents())
    use_case.storage = UploadStorage(tmp_path)
    activity = AsyncMock(side_effect=RuntimeError("activity failed"))
    use_case.activity = cast(Any, SimpleNamespace(document_uploaded=activity))

    with pytest.raises(UploadProcessingError) as exc_info:
        await use_case.upload_statement(
            context=cast(
                Any,
                SimpleNamespace(
                    workspace=SimpleNamespace(id=workspace_id),
                    user=SimpleNamespace(id=uuid4()),
                ),
            ),
            upload_file=cast(UploadFile, UploadStub()),
            account_id=account.id,
        )

    rendered = "".join(traceback.format_exception(exc_info.value))
    assert "activity failed" not in rendered
    assert "document write failed" not in rendered
    assert session.commits == 0
    assert session.rollbacks == 1
    assert activity.await_count == (1 if activity_fails else 0)
    if activity_fails:
        assert activity.await_args is not None
        assert activity.await_args.kwargs["details"].display_filename == "statement.pdf"
    assert list(tmp_path.rglob("source.pdf")) == []  # noqa: ASYNC240


def test_validate_statement_upload_accepts_pdf_and_xlsx() -> None:
    validate_statement_upload(UploadFile(file=BytesIO(b"%PDF-1.4"), filename="statement.pdf"))
    validate_statement_upload(UploadFile(file=BytesIO(b"xlsx"), filename="statement.xlsx"))


def test_validate_statement_upload_rejects_unknown_extension() -> None:
    upload = UploadFile(file=BytesIO(b"not a pdf"), filename="statement.txt")

    with pytest.raises(UploadValidationError):
        validate_statement_upload(upload)


@pytest.mark.parametrize(
    ("status", "validation_report", "expected"),
    [
        (ParseAttemptStatus.SUCCESS, None, False),
        (ParseAttemptStatus.REQUIRES_REVIEW, {"status": "valid"}, False),
        (ParseAttemptStatus.REQUIRES_REVIEW, {"status": "needs_mapping"}, True),
        (
            ParseAttemptStatus.SUCCESS,
            {"status": "valid", "source": "visual_coordinate_mapping"},
            True,
        ),
        (ParseAttemptStatus.FAILED, None, True),
        (ParseAttemptStatus.RUNNING, None, True),
    ],
)
def test_source_file_retention_policy(
    status: ParseAttemptStatus,
    validation_report: dict[str, object] | None,
    expected: bool,
) -> None:
    attempt = cast(
        ParseAttempt,
        SimpleNamespace(status=status, validation_report_json=validation_report),
    )

    assert should_retain_source_file(attempt) is expected


async def test_processed_source_is_deleted_and_marked_after_parse(tmp_path: Path) -> None:
    storage = UploadStorage(tmp_path)
    stored = await storage.save_upload(
        UploadFile(file=BytesIO(b"%PDF-1.4"), filename="statement.pdf"),
        workspace_id=uuid4(),
        document_id=uuid4(),
    )
    document = cast(
        UploadedDocument,
        SimpleNamespace(storage_key=stored.storage_key, source_file_deleted_at=None),
    )
    attempt = cast(
        ParseAttempt,
        SimpleNamespace(status=ParseAttemptStatus.SUCCESS, validation_report_json=None),
    )
    commit = AsyncMock()
    use_case = object.__new__(StatementUploadUseCase)
    use_case.storage = storage
    use_case.session = cast(Any, SimpleNamespace(commit=commit))

    await use_case._delete_processed_source(document, attempt, stored)

    assert not stored.path.exists()  # noqa: ASYNC240
    assert document.storage_key is None
    assert document.source_file_deleted_at is not None
    commit.assert_awaited_once_with()


async def test_source_deletion_failure_preserves_reference_for_cleanup(
    tmp_path: Path,
    caplog: pytest.LogCaptureFixture,
) -> None:
    storage_key = f"{uuid4()}/{uuid4()}/source.pdf"
    stored = StoredUpload(
        storage_key=storage_key,
        path=tmp_path / storage_key,
        sha256_hash="a" * 64,
        file_size_bytes=10,
    )
    document = cast(
        UploadedDocument,
        SimpleNamespace(storage_key=storage_key, source_file_deleted_at=None),
    )
    attempt = cast(
        ParseAttempt,
        SimpleNamespace(status=ParseAttemptStatus.SUCCESS, validation_report_json=None),
    )
    delete = AsyncMock(side_effect=PermissionError("sensitive path"))
    commit = AsyncMock()
    use_case = object.__new__(StatementUploadUseCase)
    use_case.storage = cast(Any, SimpleNamespace(delete_stored_upload=delete))
    use_case.session = cast(Any, SimpleNamespace(commit=commit))

    await use_case._delete_processed_source(document, attempt, stored)

    assert document.storage_key == storage_key
    assert document.source_file_deleted_at is None
    commit.assert_not_awaited()
    assert "PermissionError" in caplog.text
    assert "sensitive path" not in caplog.text


async def test_inactive_workspace_preserves_extracted_import_without_mapping_rows() -> None:
    document = SimpleNamespace(status=UploadedDocumentStatus.PARSING)
    attempt = SimpleNamespace(finished_at=None)

    class Documents:
        def __init__(self) -> None:
            self.raw_payload = None
            self.review_message = None

        async def mark_attempt_success(self, _attempt, **values):
            self.raw_payload = values

        async def mark_attempt_requires_review(self, _attempt, *, message, **_kwargs):
            self.review_message = message

        async def mark_document_status(self, target, status):
            target.status = status

    documents = Documents()
    service = StatementParseCompletionService(
        session=cast(Any, object()),
        documents=cast(Any, documents),
        statements=cast(Any, object()),
        mappings=cast(Any, object()),
        parser_registry=cast(Any, object()),
    )
    extracted = ExtractedStatement(
        text_by_page=["raw financial text"],
        tables_by_page=[
            ExtractedStatementPageTables(
                page_number=1,
                tables=[[["Дата", "Сумма"], ["2026-08-04", "-10.00"]]],
            )
        ],
        metadata={"source_format": "pdf"},
    )

    await service.preserve_inactive_workspace_attempt(
        cast(UploadedDocument, document),
        cast(ParseAttempt, attempt),
        extracted,
    )

    assert attempt.finished_at is not None
    assert documents.raw_payload == {
        "raw_text_by_page_json": ["raw financial text"],
        "raw_tables_json": [
            {"page_number": 1, "tables": [[["Дата", "Сумма"], ["2026-08-04", "-10.00"]]]}
        ],
        "metadata": {"source_format": "pdf"},
    }
    assert document.status == UploadedDocumentStatus.REQUIRES_REVIEW
    assert documents.review_message is not None
    assert "deactivated during parsing" in documents.review_message


def test_pdfplumber_extractor_preserves_raw_pages(monkeypatch: pytest.MonkeyPatch) -> None:
    class PageStub:
        def extract_text(self) -> str:
            return "synthetic statement"

        def extract_tables(self) -> list[list[list[str]]]:
            return [[["Date", "Amount"], ["2026-08-04", "-10.00"]]]

    class PdfStub:
        metadata = {"Title": "Synthetic statement"}
        pages = [PageStub()]

        def __enter__(self):
            return self

        def __exit__(self, *_args: object) -> None:
            return None

    monkeypatch.setattr(pdf_extractor_module.pdfplumber, "open", lambda _path: PdfStub())

    extracted = PdfPlumberStatementExtractor().extract(Path("synthetic.pdf"))

    assert extracted.text_by_page == ["synthetic statement"]
    assert len(extracted.tables_by_page) == len(extracted.text_by_page)
    assert all(page.page_number >= 1 for page in extracted.tables_by_page)


def test_pdf_extractor_rejects_document_over_page_limit(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    class PdfStub:
        metadata: dict[str, object] = {}
        pages = [object(), object()]

        def __enter__(self):
            return self

        def __exit__(self, *_args: object) -> None:
            return None

    monkeypatch.setattr(pdf_extractor_module.pdfplumber, "open", lambda _path: PdfStub())

    with pytest.raises(StatementResourceLimitError, match="1-page"):
        PdfPlumberStatementExtractor(StatementExtractionLimits(pdf_max_pages=1)).extract(
            Path("oversized.pdf")
        )


@pytest.mark.parametrize(
    ("limits", "text", "tables", "message"),
    [
        (StatementExtractionLimits(pdf_max_characters=6), "four", [], "character"),
        (StatementExtractionLimits(pdf_max_tables=3), "", [[[]], [[]]], "table"),
        (StatementExtractionLimits(pdf_max_cells=3), "", [[["a", "b"]]], "cell"),
    ],
)
def test_pdf_extractor_enforces_accumulated_output_limits(
    monkeypatch: pytest.MonkeyPatch,
    limits: StatementExtractionLimits,
    text: str,
    tables: list[list[list[str]]],
    message: str,
) -> None:
    class PageStub:
        def extract_text(self) -> str:
            return text

        def extract_tables(self) -> list[list[list[str]]]:
            return tables

    class PdfStub:
        metadata: dict[str, object] = {}
        pages = [PageStub(), PageStub()]

        def __enter__(self):
            return self

        def __exit__(self, *_args: object) -> None:
            return None

    monkeypatch.setattr(pdf_extractor_module.pdfplumber, "open", lambda _path: PdfStub())

    with pytest.raises(StatementResourceLimitError, match=message):
        PdfPlumberStatementExtractor(limits).extract(Path("bounded.pdf"))


def test_pdf_extractor_accepts_exact_output_boundaries(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    class PageStub:
        def extract_text(self) -> str:
            return "ab"

        def extract_tables(self) -> list[list[list[str]]]:
            return [[["cell"]]]

    class PdfStub:
        metadata: dict[str, object] = {}
        pages = [PageStub()]

        def __enter__(self):
            return self

        def __exit__(self, *_args: object) -> None:
            return None

    monkeypatch.setattr(pdf_extractor_module.pdfplumber, "open", lambda _path: PdfStub())

    extracted = PdfPlumberStatementExtractor(
        StatementExtractionLimits(pdf_max_characters=2, pdf_max_tables=1, pdf_max_cells=1)
    ).extract(Path("boundary.pdf"))

    assert extracted.text_by_page == ["ab"]


@pytest.mark.parametrize("extra_cell", [None, "over"])
def test_pdf_extractor_counts_only_non_none_cells(
    monkeypatch: pytest.MonkeyPatch,
    extra_cell: str | None,
) -> None:
    class PageStub:
        def extract_text(self) -> str:
            return ""

        def extract_tables(self) -> list[list[list[str | None]]]:
            return [[["first", None], [None, extra_cell]]]

    class PdfStub:
        metadata: dict[str, object] = {}
        pages = [PageStub()]

        def __enter__(self):
            return self

        def __exit__(self, *_args: object) -> None:
            return None

    monkeypatch.setattr(pdf_extractor_module.pdfplumber, "open", lambda _path: PdfStub())
    extractor = PdfPlumberStatementExtractor(StatementExtractionLimits(pdf_max_cells=1))

    if extra_cell is None:
        assert extractor.extract(Path("sparse.pdf")).tables_by_page[0].tables
    else:
        with pytest.raises(StatementResourceLimitError, match="cell"):
            extractor.extract(Path("sparse.pdf"))


def test_openpyxl_extractor_preserves_sheet_tables(tmp_path: Path) -> None:
    workbook_path = tmp_path / "statement.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Card"
    sheet.append(["Дата", "Описание", "Сумма"])
    sheet.append(["2026-06-01", "Coffee", -10.5])
    sheet.append([None, None, None])
    workbook.save(workbook_path)

    extracted = OpenPyxlStatementExtractor().extract(workbook_path)

    assert extracted.metadata["source_format"] == "xlsx"
    assert extracted.metadata["sheet_names"] == ["Card"]
    assert extracted.text_by_page == ["Дата\tОписание\tСумма\n2026-06-01\tCoffee\t-10.5"]
    assert extracted.tables_by_page[0].page_number == 1
    assert extracted.tables_by_page[0].tables == [
        [
            ["Дата", "Описание", "Сумма"],
            ["2026-06-01", "Coffee", "-10.5"],
        ]
    ]


def test_xlsx_extractor_rejects_large_uncompressed_archive(tmp_path: Path) -> None:
    workbook_path = tmp_path / "oversized.xlsx"
    with ZipFile(workbook_path, "w", compression=ZIP_DEFLATED) as archive:
        archive.writestr("xl/worksheets/sheet1.xml", "x" * 101)

    with pytest.raises(StatementResourceLimitError, match="uncompressed size"):
        OpenPyxlStatementExtractor(
            StatementExtractionLimits(xlsx_max_uncompressed_bytes=100)
        ).extract(workbook_path)


@pytest.mark.parametrize(
    ("limits", "cell", "message"),
    [
        (StatementExtractionLimits(xlsx_max_rows_per_sheet=1), "A2", "row"),
        (StatementExtractionLimits(xlsx_max_columns_per_sheet=1), "B1", "column"),
        (StatementExtractionLimits(xlsx_max_cells=3), "B2", "total cell"),
    ],
)
def test_xlsx_extractor_enforces_sheet_dimensions(
    tmp_path: Path,
    limits: StatementExtractionLimits,
    cell: str,
    message: str,
) -> None:
    workbook_path = tmp_path / "oversized.xlsx"
    workbook = Workbook()
    workbook.active[cell] = "value"
    workbook.save(workbook_path)

    with pytest.raises(StatementResourceLimitError, match=message):
        OpenPyxlStatementExtractor(limits).extract(workbook_path)


def test_xlsx_extractor_enforces_sheet_count(tmp_path: Path) -> None:
    workbook_path = tmp_path / "oversized.xlsx"
    workbook = Workbook()
    workbook.create_sheet("Second")
    workbook.save(workbook_path)

    with pytest.raises(StatementResourceLimitError, match="1-sheet"):
        OpenPyxlStatementExtractor(StatementExtractionLimits(xlsx_max_sheets=1)).extract(
            workbook_path
        )


async def test_resource_limit_failure_preserves_document_and_failed_attempt() -> None:
    document = SimpleNamespace(status=UploadedDocumentStatus.PARSING)
    attempt = SimpleNamespace(finished_at=None)

    class Documents:
        def __init__(self) -> None:
            self.error_code: str | None = None
            self.error_message: str | None = None

        async def mark_attempt_failed(
            self,
            _attempt: object,
            *,
            error_code: str,
            error_message: str,
        ) -> None:
            self.error_code = error_code
            self.error_message = error_message

        async def mark_document_status(
            self,
            target: Any,
            status: UploadedDocumentStatus,
        ) -> None:
            target.status = status

    documents = Documents()
    await record_failed_parse_attempt(
        cast(Any, documents),
        cast(UploadedDocument, document),
        cast(ParseAttempt, attempt),
        StatementResourceLimitError("XLSX exceeds configured limits."),
    )

    assert attempt.finished_at is not None
    assert documents.error_code == "StatementResourceLimitError"
    assert documents.error_message == "StatementResourceLimitError"
    assert document.status is UploadedDocumentStatus.FAILED_TO_PARSE


async def test_completion_failure_rolls_back_then_commits_terminal_attempt_with_raw() -> None:
    workspace_id = uuid4()
    document = SimpleNamespace(id=uuid4(), status=UploadedDocumentStatus.PARSING)
    attempt = SimpleNamespace(id=uuid4(), status=ParseAttemptStatus.RUNNING, finished_at=None)

    class Documents:
        async def get_document_for_workspace(self, requested_workspace_id, document_id):
            assert (requested_workspace_id, document_id) == (workspace_id, document.id)
            return document

        async def get_parse_attempt_for_workspace(
            self, requested_workspace_id, document_id, attempt_id
        ):
            assert (requested_workspace_id, document_id, attempt_id) == (
                workspace_id,
                document.id,
                attempt.id,
            )
            return attempt

        async def store_attempt_extracted_raw(self, target, **payload):
            target.raw_payload = payload

        async def mark_attempt_failed(self, target, *, error_code, error_message):
            target.status = ParseAttemptStatus.FAILED
            target.error_code = error_code
            target.error_message = error_message

        async def mark_document_status(self, target, status):
            target.status = status

    session = SimpleNamespace(rollback=AsyncMock(), commit=AsyncMock())
    use_case = object.__new__(StatementUploadUseCase)
    use_case.session = cast(Any, session)
    use_case.documents = cast(Any, Documents())
    extracted = ExtractedStatement(
        text_by_page=["bounded raw"],
        tables_by_page=[ExtractedStatementPageTables(page_number=1, tables=[])],
        metadata={"source_format": "pdf"},
    )

    await use_case._commit_terminal_failure(
        workspace_id=workspace_id,
        document_id=document.id,
        attempt_id=attempt.id,
        error=RuntimeError("private /path/source.pdf contents"),
        extracted=extracted,
    )

    session.rollback.assert_awaited_once()
    session.commit.assert_awaited_once()
    assert attempt.status is ParseAttemptStatus.FAILED
    assert document.status is UploadedDocumentStatus.FAILED_TO_PARSE
    assert attempt.raw_payload["raw_text_by_page_json"] == ["bounded raw"]
    assert attempt.error_message == "RuntimeError"
    assert "/path" not in attempt.error_message


@pytest.mark.parametrize(
    ("stage", "failure", "expected_error"),
    [
        pytest.param("extractor", ValueError("extractor failed"), None, id="extractor"),
        pytest.param(
            "completion", RuntimeError("analyzer failed"), UploadProcessingError, id="analyzer"
        ),
        pytest.param(
            "completion", LookupError("mapping failed"), UploadProcessingError, id="mapping"
        ),
        pytest.param(
            "extractor",
            RuntimeError("RAW_MARKER /private/statement.pdf"),
            UploadProcessingError,
            id="unexpected",
        ),
        pytest.param(
            "extractor", asyncio.CancelledError(), asyncio.CancelledError, id="cancellation"
        ),
    ],
)
async def test_whole_upload_failure_commits_terminal_attempt(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
    stage: str,
    failure: BaseException,
    expected_error: type[BaseException] | None,
) -> None:
    use_case, context, document, attempt, session = _whole_upload_harness(
        monkeypatch,
        tmp_path,
        extract_error=failure if stage == "extractor" else None,
        completion_error=failure if stage == "completion" else None,
    )

    if expected_error is None:
        result = await use_case.upload_statement(
            context=context,
            upload_file=UploadFile(file=BytesIO(b"%PDF-1.4"), filename="statement.pdf"),
            account_id=document.account_id,
        )
        assert result.document_status is UploadedDocumentStatus.FAILED_TO_PARSE
    else:
        with pytest.raises(expected_error) as raised:
            await use_case.upload_statement(
                context=context,
                upload_file=UploadFile(file=BytesIO(b"%PDF-1.4"), filename="statement.pdf"),
                account_id=document.account_id,
            )
        if expected_error is UploadProcessingError:
            rendered = "".join(traceback.format_exception(raised.value))
            assert "RAW_MARKER" not in rendered
            assert "/private/statement.pdf" not in rendered

    assert attempt.status is ParseAttemptStatus.FAILED
    assert document.status is UploadedDocumentStatus.FAILED_TO_PARSE
    assert session.commit.await_count == 3
    session.rollback.assert_awaited_once()


async def test_cancelled_upload_bounds_hung_terminal_cleanup(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    use_case, context, document, _attempt, session = _whole_upload_harness(
        monkeypatch,
        tmp_path,
        extract_error=asyncio.CancelledError(),
    )

    async def hang_rollback() -> None:
        await asyncio.Event().wait()

    session.rollback = AsyncMock(side_effect=hang_rollback)
    monkeypatch.setattr(upload_module, "FAILURE_CLEANUP_TIMEOUT_SECONDS", 0.01)

    with pytest.raises(asyncio.CancelledError):
        await asyncio.wait_for(
            use_case.upload_statement(
                context=context,
                upload_file=UploadFile(file=BytesIO(b"%PDF-1.4"), filename="statement.pdf"),
                account_id=document.account_id,
            ),
            timeout=0.2,
        )


@pytest.mark.parametrize(
    ("commit_error", "expected_error"),
    [
        pytest.param(asyncio.CancelledError(), asyncio.CancelledError, id="cancellation"),
        pytest.param(RuntimeError("commit result unknown"), UploadProcessingError, id="error"),
    ],
)
async def test_attempt_commit_boundary_recovers_persisted_running_attempt(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
    commit_error: BaseException,
    expected_error: type[BaseException],
) -> None:
    use_case, context, document, attempt, session = _whole_upload_harness(
        monkeypatch,
        tmp_path,
    )
    document.status = UploadedDocumentStatus.UPLOADED
    commit_calls = 0

    async def commit_with_unknown_second_result() -> None:
        nonlocal commit_calls
        commit_calls += 1
        if commit_calls == 2:
            raise commit_error

    session.commit = AsyncMock(side_effect=commit_with_unknown_second_result)

    with pytest.raises(expected_error):
        await use_case.upload_statement(
            context=context,
            upload_file=UploadFile(file=BytesIO(b"%PDF-1.4"), filename="statement.pdf"),
            account_id=document.account_id,
        )

    assert commit_calls == 3
    assert attempt.status is ParseAttemptStatus.FAILED
    assert document.status is UploadedDocumentStatus.FAILED_TO_PARSE
    session.rollback.assert_awaited_once()


async def test_attempt_commit_boundary_does_not_create_missing_failure() -> None:
    session = SimpleNamespace(rollback=AsyncMock(), commit=AsyncMock())
    documents = SimpleNamespace(
        get_parse_attempt_for_workspace=AsyncMock(return_value=None),
        get_document_for_workspace=AsyncMock(),
    )
    use_case = object.__new__(StatementUploadUseCase)
    use_case.session = cast(Any, session)
    use_case.documents = cast(Any, documents)

    await use_case._commit_terminal_failure_if_persisted(
        workspace_id=uuid4(),
        document_id=uuid4(),
        attempt_id=uuid4(),
        error=RuntimeError("commit failed before persistence"),
    )

    session.rollback.assert_awaited_once()
    session.commit.assert_not_awaited()
    documents.get_document_for_workspace.assert_not_awaited()


async def test_unexpected_storage_failure_is_sanitized_and_removes_partial_file(
    tmp_path: Path,
) -> None:
    marker = "STORAGE_PRIVATE_MARKER /private/source.pdf"

    class FailingUploadStream(BytesIO):
        def read(self, size: int | None = -1, /) -> bytes:
            if self.tell() == len(self.getvalue()):
                raise RuntimeError(marker)
            return super().read(size)

    workspace_id = uuid4()
    account_id = uuid4()
    session = SimpleNamespace(rollback=AsyncMock(), commit=AsyncMock())
    use_case = object.__new__(StatementUploadUseCase)
    use_case.session = cast(Any, session)
    use_case.settings = SimpleNamespace(statement_upload_max_bytes=1024)
    use_case.accounts = SimpleNamespace(
        get_import_account=AsyncMock(return_value=SimpleNamespace(id=account_id, currency="RUB"))
    )
    use_case.storage = UploadStorage(tmp_path)
    context = SimpleNamespace(workspace=SimpleNamespace(id=workspace_id))

    with pytest.raises(UploadProcessingError) as exc_info:
        await use_case.upload_statement(
            context=context,
            upload_file=UploadFile(
                file=FailingUploadStream(b"%PDF-1.4 partial"),
                filename="statement.pdf",
            ),
            account_id=account_id,
        )

    rendered = "".join(traceback.format_exception(exc_info.value))
    assert exc_info.value.__cause__ is None
    assert marker not in rendered
    assert "/private/source.pdf" not in rendered
    assert list(tmp_path.rglob("source.pdf")) == []  # noqa: ASYNC240
    session.rollback.assert_awaited_once()


async def test_unexpected_initial_document_failure_is_sanitized_and_cleans_source(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    marker = "INITIAL_DB_PRIVATE_MARKER sql-params"
    use_case, context, document, _attempt, session = _whole_upload_harness(
        monkeypatch,
        tmp_path,
    )
    cast(Any, use_case)._create_document = AsyncMock(side_effect=RuntimeError(marker))

    with pytest.raises(UploadProcessingError) as exc_info:
        await use_case.upload_statement(
            context=context,
            upload_file=UploadFile(file=BytesIO(b"%PDF-1.4"), filename="statement.pdf"),
            account_id=document.account_id,
        )

    rendered = "".join(traceback.format_exception(exc_info.value))
    assert exc_info.value.__cause__ is None
    assert marker not in rendered
    session.rollback.assert_awaited_once()
    cast(Any, use_case.storage.delete_stored_upload).assert_awaited_once()


async def test_initial_commit_reconciliation_is_bounded(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    async def hang_reconciliation(**_kwargs: object) -> None:
        await asyncio.Event().wait()

    use_case = object.__new__(StatementUploadUseCase)
    cast(Any, use_case)._reconcile_initial_commit = hang_reconciliation
    monkeypatch.setattr(upload_module, "FAILURE_CLEANUP_TIMEOUT_SECONDS", 0.01)

    with pytest.raises(UploadProcessingError):
        await asyncio.wait_for(
            use_case._reconcile_initial_commit_bounded(
                workspace_id=uuid4(),
                document_id=uuid4(),
                attempt_id=uuid4(),
                stored_upload=StoredUpload(
                    storage_key="workspace/document/source.pdf",
                    path=Path("source.pdf"),
                    sha256_hash="a" * 64,
                    file_size_bytes=8,
                ),
                error=RuntimeError("ambiguous commit"),
            ),
            timeout=0.2,
        )


def _whole_upload_harness(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
    *,
    extract_error: BaseException | None = None,
    completion_error: BaseException | None = None,
) -> tuple[StatementUploadUseCase, Any, Any, Any, Any]:
    workspace_id = uuid4()
    account_id = uuid4()
    document = SimpleNamespace(
        id=uuid4(),
        account_id=account_id,
        status=UploadedDocumentStatus.PARSING,
        original_filename="statement.pdf",
    )
    attempt = SimpleNamespace(id=uuid4(), status=ParseAttemptStatus.RUNNING, finished_at=None)

    class Documents:
        async def get_document_for_workspace(self, _workspace_id, _document_id):
            return document

        async def get_parse_attempt_for_workspace(self, *_args):
            return attempt

        async def store_attempt_extracted_raw(self, target, **payload):
            target.raw_payload = payload

        async def mark_attempt_failed(self, target, *, error_code, error_message):
            target.status = ParseAttemptStatus.FAILED
            target.error_code = error_code
            target.error_message = error_message

        async def mark_document_status(self, target, status):
            target.status = status

    async def extract(_path: Path) -> ExtractedStatement:
        if extract_error is not None:
            raise extract_error
        return ExtractedStatement(text_by_page=["bounded"], tables_by_page=[])

    async def complete(*_args: object, **_kwargs: object) -> None:
        if completion_error is not None:
            raise completion_error

    async def create_attempt(*_args: object, **_kwargs: object):
        return attempt

    monkeypatch.setattr(upload_module, "create_running_parse_attempt", create_attempt)
    session = SimpleNamespace(commit=AsyncMock(), rollback=AsyncMock())
    use_case = object.__new__(StatementUploadUseCase)
    use_case.session = cast(Any, session)
    use_case.settings = SimpleNamespace(statement_upload_max_bytes=1024)
    use_case.accounts = SimpleNamespace(
        get_import_account=AsyncMock(return_value=SimpleNamespace(id=account_id, currency="RUB"))
    )
    use_case.documents = cast(Any, Documents())
    use_case.storage = SimpleNamespace(
        save_upload=AsyncMock(
            return_value=StoredUpload(
                storage_key=f"{workspace_id}/{document.id}/source.pdf",
                path=tmp_path / "source.pdf",
                sha256_hash="a" * 64,
                file_size_bytes=8,
            )
        ),
        delete_stored_upload=AsyncMock(),
    )
    use_case.activity = SimpleNamespace(document_uploaded=AsyncMock())
    use_case.workspaces = SimpleNamespace(
        lock_for_update=AsyncMock(return_value=SimpleNamespace(is_active=True))
    )
    use_case.parse_completion = SimpleNamespace(complete_successful_attempt=complete)
    use_case._create_document = AsyncMock(return_value=document)
    use_case._extract_statement = extract
    context = SimpleNamespace(
        workspace=SimpleNamespace(id=workspace_id),
        user=SimpleNamespace(id=uuid4()),
    )
    return use_case, context, document, attempt, session


def test_statement_extractor_resolver_selects_extractor_by_extension(tmp_path: Path) -> None:
    workbook_path = tmp_path / "statement.xlsx"
    workbook = Workbook()
    workbook.active.append(["Date", "Amount"])
    workbook.save(workbook_path)

    extracted = StatementExtractorResolver().extract(workbook_path)

    assert extracted.metadata["source_format"] == "xlsx"


def test_reviewable_raw_transaction_statuses_include_normalized_rows() -> None:
    assert RawTransactionStatus.NORMALIZED in REVIEW_QUEUE_STATUSES
    assert RawTransactionStatus.SUGGESTED in REVIEW_QUEUE_STATUSES
    assert RawTransactionStatus.MATCHED in REVIEW_QUEUE_STATUSES
    assert RawTransactionStatus.CONFIRMED not in REVIEW_QUEUE_STATUSES
    assert RawTransactionStatus.IGNORED not in REVIEW_QUEUE_STATUSES
    assert RawTransactionStatus.DUPLICATE not in REVIEW_QUEUE_STATUSES


def test_raw_transaction_payload_keeps_provenance_without_full_source_rows() -> None:
    draft = RawTransactionDraft(
        row_index=0,
        status=RawTransactionStatus.NORMALIZED,
        raw_payload={
            "bank_code": "test",
            "source_row_id": "stable:1",
            "cells": ["full", "private", "row"],
            "raw_row": "full private source line",
        },
        operation_date_raw="2026-08-26",
        posting_date_raw=None,
        description_raw="Description retained for review",
        amount_raw="10.00",
        currency_raw="RUB",
        balance_after_raw=None,
        account_hint_raw=None,
        account_id=uuid4(),
        operation_date=date(2026, 8, 26),
        posting_date=date(2026, 8, 26),
        description_normalized="Description retained for review",
        amount=Decimal("10.00"),
        currency="RUB",
        balance_after=None,
        dedupe_hash="hash",
        confidence_score=Decimal("1"),
        normalization_error=None,
    )

    raw = RawTransactionMapper.from_draft(
        draft,
        workspace_id=uuid4(),
        uploaded_document_id=uuid4(),
        parse_attempt_id=uuid4(),
    )

    assert raw.raw_payload == {"bank_code": "test", "source_row_id": "stable:1"}
    assert raw.description_raw == "Description retained for review"


async def test_import_account_masks_archived_non_debt_account() -> None:
    workspace_id = uuid4()
    account_id = uuid4()
    resolver = object.__new__(LedgerReferenceResolver)
    resolver.accounts = cast(
        Any,
        SimpleNamespace(
            get_for_workspace=AsyncMock(
                return_value=SimpleNamespace(
                    id=account_id,
                    type=AccountType.CASH,
                    is_active=False,
                )
            )
        ),
    )

    with pytest.raises(LedgerPostingError, match="archived"):
        await resolver.get_import_account(workspace_id, account_id)


async def test_duplicate_candidate_query_is_workspace_and_document_scoped() -> None:
    workspace_id = uuid4()
    document_id = uuid4()
    account_id = uuid4()

    class ResultStub:
        def scalars(self) -> Any:
            return self

        def all(self) -> list[object]:
            return []

    class SessionStub:
        statement: object | None = None

        async def execute(self, statement: object) -> ResultStub:
            self.statement = statement
            return ResultStub()

    session = SessionStub()
    await ImportReviewRepository(cast(Any, session)).list_possible_duplicate_candidates(
        workspace_id=workspace_id,
        fingerprints={(account_id, date(2026, 7, 20), Decimal("-1250.50"), "RUB")},
        exclude_document_id=document_id,
    )

    assert session.statement is not None
    compiled = cast(Any, session.statement).compile()
    assert workspace_id in compiled.params.values()
    assert document_id in compiled.params.values()
    assert "raw_transactions.workspace_id" in str(compiled)
    assert "raw_transactions.uploaded_document_id !=" in str(compiled)


def test_possible_duplicate_fingerprint_requires_normalized_fields() -> None:
    account_id = uuid4()
    raw_transaction = raw_transaction_from_values(
        account_id=account_id,
        amount=Decimal("10.00"),
    )

    assert possible_duplicate_fingerprint(raw_transaction) == (
        account_id,
        parse_bank_date("29.05.2026"),
        Decimal("10.00"),
        "RUB",
    )
    raw_transaction.amount = None
    assert possible_duplicate_fingerprint(raw_transaction) is None


def raw_transaction_from_values(
    *,
    account_id: UUID | None = None,
    amount: Decimal | None = Decimal("10.00"),
) -> RawTransaction:
    return RawTransaction(
        workspace_id=uuid4(),
        uploaded_document_id=uuid4(),
        parse_attempt_id=uuid4(),
        row_index=0,
        status=RawTransactionStatus.NORMALIZED,
        raw_payload={},
        account_id=account_id,
        operation_date=parse_bank_date("29.05.2026"),
        amount=amount,
        currency="RUB",
        normalization_error=None,
    )
