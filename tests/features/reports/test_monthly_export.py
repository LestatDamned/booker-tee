from datetime import UTC, date, datetime
from decimal import Decimal
from io import BytesIO
from typing import Any, cast
from unittest.mock import AsyncMock
from uuid import uuid4

import pytest
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.features.ledger.domain.types import OperationSource, OperationType
from app.features.reports.application.monthly_export import (
    MAX_MONTHLY_EXPORT_ENTRIES,
    InvalidReportMonthError,
    MonthlyExportTooLargeError,
    MonthlyReportData,
    MonthlyReportReader,
    month_date_range,
)
from app.features.reports.application.overview import (
    ReportBalanceSummary,
    ReportingFilterOptions,
    ReportingFilters,
    ReportingOverview,
)
from app.features.reports.monthly_report_xlsx import build_monthly_report_xlsx
from app.features.reports.repository import (
    ReportAccountBalanceRow,
    ReportMoneySummaryRow,
    ReportOperationEntryRow,
    ReportsRepository,
    ReportUncategorizedPage,
)


@pytest.mark.parametrize(
    ("month", "expected"),
    [
        ("2024-02", (date(2024, 2, 1), date(2024, 2, 29))),
        ("2026-12", (date(2026, 12, 1), date(2026, 12, 31))),
    ],
)
def test_month_date_range(month: str, expected: tuple[date, date]) -> None:
    assert month_date_range(month) == expected


@pytest.mark.parametrize("month", ["2026-8", "08.2026", "2026-13", ""])
def test_month_date_range_rejects_invalid_values(month: str) -> None:
    with pytest.raises(InvalidReportMonthError):
        month_date_range(month)


async def test_monthly_reader_reuses_overview_and_enforces_entry_limit() -> None:
    workspace_id = uuid4()
    requested_filters = ReportingFilters(
        date_from=date(2026, 7, 1),
        date_to=date(2026, 7, 31),
        currency="rub",
    )
    applied_filters = ReportingFilters(
        date_from=requested_filters.date_from,
        date_to=requested_filters.date_to,
        currency="RUB",
    )
    overview_reader = AsyncMock()
    overview_reader.read.return_value = reporting_overview(applied_filters)
    repository = AsyncMock()
    repository.list_operation_entries.return_value = [operation_entry()] * (
        MAX_MONTHLY_EXPORT_ENTRIES + 1
    )
    reader = MonthlyReportReader(cast(Any, overview_reader), cast(Any, repository))

    with pytest.raises(MonthlyExportTooLargeError):
        await reader.read(
            workspace_id=workspace_id,
            workspace_name="Личные финансы",
            default_currency="RUB",
            month="2026-07",
            currency="rub",
        )

    overview_reader.read.assert_awaited_once_with(
        workspace_id=workspace_id,
        default_currency="RUB",
        filters=requested_filters,
    )
    repository.list_operation_entries.assert_awaited_once_with(
        workspace_id=workspace_id,
        filters=applied_filters,
        limit=MAX_MONTHLY_EXPORT_ENTRIES + 1,
    )


async def test_operation_entry_projection_is_bounded_and_preserves_provenance() -> None:
    workspace_id = uuid4()
    operation_id = uuid4()
    session = EntrySessionCapture(
        entry_rows=[
            (
                operation_id,
                1,
                date(2026, 7, 15),
                None,
                OperationType.EXPENSE,
                OperationSource.BANK_PDF,
                "Основной",
                Decimal("-250.00"),
                "RUB",
                True,
                "Без категории",
                None,
                None,
            )
        ],
        provenance_rows=[
            (operation_id, "statement.pdf", 4),
            (operation_id, "statement.pdf", 4),
            (operation_id, "statement-2.xlsx", 8),
        ],
    )

    rows = await ReportsRepository(cast(AsyncSession, session)).list_operation_entries(
        workspace_id=workspace_id,
        filters=ReportingFilters(
            date_from=date(2026, 7, 1),
            date_to=date(2026, 7, 31),
            currency="RUB",
        ),
        limit=10_001,
    )

    assert rows == [
        ReportOperationEntryRow(
            operation_id=operation_id,
            entry_order=1,
            operation_date=date(2026, 7, 15),
            posting_date=None,
            operation_type=OperationType.EXPENSE,
            source=OperationSource.BANK_PDF,
            account_name="Основной",
            amount=Decimal("-250.00"),
            currency="RUB",
            affects_profit=True,
            category_name="Без категории",
            property_name=None,
            description="Без описания",
            import_documents="statement.pdf; statement-2.xlsx",
            import_rows="4; 8",
        )
    ]
    entry_query, provenance_query = session.queries
    entry_sql = str(entry_query)
    assert "operations.status" in entry_sql
    assert "operations.operation_date >=" in entry_sql
    assert "operations.operation_date <=" in entry_sql
    assert "operations.affects_profit IS true" not in entry_sql
    assert workspace_id in entry_query.compile().params.values()
    assert 10_001 in entry_query.compile().params.values()
    assert "raw_transactions.workspace_id" in str(provenance_query)


def test_xlsx_builder_writes_stable_typed_and_safe_workbook() -> None:
    operation_id = uuid4()
    overview = reporting_overview(
        ReportingFilters(
            date_from=date(2026, 7, 1),
            date_to=date(2026, 7, 31),
            currency="RUB",
        )
    )
    entries = [
        ReportOperationEntryRow(
            operation_id=operation_id,
            entry_order=1,
            operation_date=date(2026, 7, 15),
            posting_date=None,
            operation_type=OperationType.TRANSFER,
            source=OperationSource.BANK_PDF,
            account_name="Основной",
            amount=Decimal("-250.00"),
            currency="RUB",
            affects_profit=False,
            category_name=None,
            property_name=None,
            description='=HYPERLINK("https://example.test")',
            import_documents="statement.xlsx",
            import_rows="4",
        ),
        ReportOperationEntryRow(
            operation_id=operation_id,
            entry_order=2,
            operation_date=date(2026, 7, 15),
            posting_date=None,
            operation_type=OperationType.TRANSFER,
            source=OperationSource.BANK_PDF,
            account_name="Наличные",
            amount=Decimal("250.00"),
            currency="RUB",
            affects_profit=False,
            category_name=None,
            property_name=None,
            description="Перевод",
            import_documents="statement.xlsx",
            import_rows="4",
        ),
    ]
    payload = build_monthly_report_xlsx(
        MonthlyReportData(
            workspace_name="Личные финансы",
            month="2026-07",
            generated_at=datetime(2026, 8, 1, tzinfo=UTC),
            overview=overview,
            entries=entries,
        )
    )

    workbook = load_workbook(BytesIO(payload), data_only=False)
    assert workbook.sheetnames == ["Итоги", "Счета", "Категории", "Объекты", "Операции"]
    assert workbook["Итоги"]["B8"].value == 100
    operations = workbook["Операции"]
    assert operations.max_row == 3
    assert operations["A2"].is_date is True
    assert operations["F2"].data_type == "n"
    assert operations["F2"].number_format.startswith("#,##0.00")
    assert operations["K2"].value == '=HYPERLINK("https://example.test")'
    assert operations["K2"].data_type == "s"
    assert operations["N2"].value == operations["N3"].value == str(operation_id)
    assert operations.auto_filter.ref == "A1:O3"


class QueryResult:
    def __init__(self, rows: list[tuple[Any, ...]]) -> None:
        self.rows = rows

    def all(self) -> list[tuple[Any, ...]]:
        return self.rows


class EntrySessionCapture:
    def __init__(
        self,
        *,
        entry_rows: list[tuple[Any, ...]],
        provenance_rows: list[tuple[Any, ...]],
    ) -> None:
        self.results = [entry_rows, provenance_rows]
        self.queries: list[Any] = []

    async def execute(self, query: Any) -> QueryResult:
        self.queries.append(query)
        return QueryResult(self.results[len(self.queries) - 1])


def reporting_overview(filters: ReportingFilters) -> ReportingOverview:
    currency = filters.currency or "RUB"
    return ReportingOverview(
        filters=filters,
        filter_options=ReportingFilterOptions(
            accounts=[], categories=[], properties=[], currencies=[currency]
        ),
        summary=ReportMoneySummaryRow(
            currency,
            Decimal("100.00"),
            Decimal("40.00"),
            Decimal("60.00"),
        ),
        balance_summary=ReportBalanceSummary(
            currency,
            Decimal("1000.00"),
            Decimal("1060.00"),
            Decimal("60.00"),
        ),
        account_balances=[
            ReportAccountBalanceRow(
                uuid4(),
                "Основной",
                currency,
                Decimal("1000.00"),
                Decimal("1060.00"),
                Decimal("60.00"),
                True,
            )
        ],
        categories=[],
        properties=[],
        balance_as_of=filters.date_to,
        next_review_document_id=None,
        uncategorized=ReportUncategorizedPage(items=[], page=1, page_size=10, total=0),
    )


def operation_entry() -> ReportOperationEntryRow:
    return ReportOperationEntryRow(
        operation_id=uuid4(),
        entry_order=1,
        operation_date=date(2026, 7, 15),
        posting_date=None,
        operation_type=OperationType.EXPENSE,
        source=OperationSource.MANUAL,
        account_name="Основной",
        amount=Decimal("-10.00"),
        currency="RUB",
        affects_profit=True,
        category_name=None,
        property_name=None,
        description="Кофе",
        import_documents=None,
        import_rows=None,
    )
