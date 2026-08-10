from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from app.features.ledger.domain.types import OperationSource, OperationType
from app.features.reports.application.monthly_export import MonthlyReportData

MONEY_FORMAT = "#,##0.00;[Red]-#,##0.00"
DATE_FORMAT = "DD.MM.YYYY"
HEADER_FILL = PatternFill("solid", fgColor="243247")
HEADER_FONT = Font(color="FFFFFF", bold=True)

SOURCE_LABELS = {
    OperationSource.MANUAL: "Ручная",
    OperationSource.BANK_PDF: "Импорт",
    OperationSource.DEBT: "Долг",
    OperationSource.SYSTEM: "Системная",
}
TYPE_LABELS = {
    OperationType.INCOME: "Доход",
    OperationType.EXPENSE: "Расход",
    OperationType.TRANSFER: "Перевод",
    OperationType.ADJUSTMENT: "Корректировка",
}


def build_monthly_report_xlsx(report: MonthlyReportData) -> bytes:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Итоги"
    _write_summary(summary, report)
    _write_accounts(workbook.create_sheet("Счета"), report)
    _write_categories(workbook.create_sheet("Категории"), report)
    _write_properties(workbook.create_sheet("Объекты"), report)
    _write_entries(workbook.create_sheet("Операции"), report)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _write_summary(sheet: Worksheet, report: MonthlyReportData) -> None:
    overview = report.overview
    rows = [
        ("Месячный финансовый отчёт", None, None),
        ("Workspace", report.workspace_name, None),
        ("Месяц", report.month, None),
        ("Валюта", overview.summary.currency, None),
        ("Сформирован", report.generated_at.replace(tzinfo=None), None),
        (None, None, None),
        ("Показатель", "Значение", "Валюта"),
        ("Доходы", overview.summary.income, overview.summary.currency),
        ("Расходы", overview.summary.expense, overview.summary.currency),
        ("Итог периода", overview.summary.profit, overview.summary.currency),
        (
            "Результат",
            (
                "Положительный"
                if overview.summary.profit > 0
                else "Отрицательный"
                if overview.summary.profit < 0
                else "Нулевой"
            ),
            None,
        ),
        ("Остаток на начало", overview.balance_summary.opening_balance, overview.summary.currency),
        ("Остаток на конец", overview.balance_summary.closing_balance, overview.summary.currency),
        ("Изменение остатка", overview.balance_summary.balance_change, overview.summary.currency),
        (
            "Методика",
            "Переводы видны в операциях, но не входят в доходы, расходы и итог.",
            None,
        ),
        (
            "Данные на проверке",
            "Есть" if overview.next_review_document_id else "Нет",
            None,
        ),
    ]
    for row in rows:
        _append(sheet, row)
    _style_header(sheet, 7)
    for row in (8, 9, 10, 12, 13, 14):
        sheet.cell(row, 2).number_format = MONEY_FORMAT
    sheet.cell(5, 2).number_format = "DD.MM.YYYY HH:MM"
    sheet.freeze_panes = "A8"
    _set_widths(sheet, (30, 76, 12))


def _write_accounts(sheet: Worksheet, report: MonthlyReportData) -> None:
    _append(sheet, ("Счёт", "Статус", "Валюта", "На начало", "На конец", "Изменение"))
    for row in report.overview.account_balances:
        _append(
            sheet,
            (
                row.name,
                "Активен" if row.is_active else "Архив",
                row.currency,
                row.opening_balance,
                row.closing_balance,
                row.balance_change,
            ),
        )
    _finish_table(sheet, money_columns=(4, 5, 6), widths=(28, 12, 10, 16, 16, 16))


def _write_categories(sheet: Worksheet, report: MonthlyReportData) -> None:
    _append(sheet, ("Категория", "Статус", "Доходы", "Расходы", "Итог"))
    for row in report.overview.categories:
        _append(
            sheet,
            (
                row.name,
                "Активна" if row.is_active else "Архив",
                row.income,
                row.expense,
                row.profit,
            ),
        )
    _finish_table(sheet, money_columns=(3, 4, 5), widths=(32, 12, 16, 16, 16))


def _write_properties(sheet: Worksheet, report: MonthlyReportData) -> None:
    _append(sheet, ("Объект", "Статус", "Доходы", "Расходы", "Итог"))
    for row in report.overview.properties:
        _append(
            sheet,
            (
                row.name,
                "Активен" if row.is_active else "Архив",
                row.income,
                row.expense,
                row.profit,
            ),
        )
    _finish_table(sheet, money_columns=(3, 4, 5), widths=(32, 12, 16, 16, 16))


def _write_entries(sheet: Worksheet, report: MonthlyReportData) -> None:
    _append(
        sheet,
        (
            "Дата операции",
            "Дата проводки",
            "Тип",
            "Источник",
            "Счёт",
            "Движение",
            "Валюта",
            "Влияет на результат",
            "Категория",
            "Объект",
            "Описание",
            "Документ импорта",
            "Строка импорта",
            "ID операции",
            "Номер движения",
        ),
    )
    for row in report.entries:
        _append(
            sheet,
            (
                row.operation_date,
                row.posting_date,
                TYPE_LABELS[row.operation_type],
                SOURCE_LABELS[row.source],
                row.account_name,
                row.amount,
                row.currency,
                "Да" if row.affects_profit else "Нет",
                row.category_name,
                row.property_name,
                row.description,
                row.import_documents,
                row.import_rows,
                str(row.operation_id),
                row.entry_order,
            ),
        )
    _finish_table(
        sheet,
        money_columns=(6,),
        date_columns=(1, 2),
        widths=(14, 14, 16, 14, 28, 16, 10, 18, 26, 26, 44, 28, 16, 38, 14),
    )


def _append(sheet: Worksheet, values: tuple[Any, ...]) -> None:
    row = 1 if sheet.max_row == 1 and sheet.cell(1, 1).value is None else sheet.max_row + 1
    for column, value in enumerate(values, start=1):
        cell = sheet.cell(row, column)
        cell.value = value
        if isinstance(value, str):
            cell.data_type = "s"


def _finish_table(
    sheet: Worksheet,
    *,
    money_columns: tuple[int, ...],
    widths: tuple[int, ...],
    date_columns: tuple[int, ...] = (),
) -> None:
    _style_header(sheet, 1)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for row in range(2, sheet.max_row + 1):
        for column in money_columns:
            sheet.cell(row, column).number_format = MONEY_FORMAT
        for column in date_columns:
            sheet.cell(row, column).number_format = DATE_FORMAT
    _set_widths(sheet, widths)


def _style_header(sheet: Worksheet, row: int) -> None:
    for cell in sheet[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")


def _set_widths(sheet: Worksheet, widths: tuple[int, ...]) -> None:
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[sheet.cell(1, column).column_letter].width = width
