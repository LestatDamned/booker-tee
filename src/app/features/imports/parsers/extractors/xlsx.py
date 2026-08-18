from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from app.features.imports.parsers.extractors.dto import (
    ExtractedStatement,
    ExtractedStatementPageTables,
)
from app.features.imports.parsers.extractors.limits import (
    StatementExtractionLimits,
    StatementResourceLimitError,
)


class OpenPyxlStatementExtractor:
    parser_name = "openpyxl_raw_extractor"
    parser_version = "0.1"

    def __init__(self, limits: StatementExtractionLimits | None = None) -> None:
        self.limits = limits or StatementExtractionLimits()

    def extract(self, file_path: Path) -> ExtractedStatement:
        validate_xlsx_archive(file_path, self.limits.xlsx_max_uncompressed_bytes)
        workbook = load_workbook(
            file_path,
            read_only=True,
            data_only=True,
        )
        try:
            if len(workbook.sheetnames) > self.limits.xlsx_max_sheets:
                raise StatementResourceLimitError(
                    f"XLSX exceeds the {self.limits.xlsx_max_sheets}-sheet extraction limit."
                )
            text_by_page: list[str] = []
            tables_by_page: list[ExtractedStatementPageTables] = []
            total_cells = 0
            for sheet_index, worksheet in enumerate(workbook.worksheets, start=1):
                rows = worksheet.max_row or 0
                columns = worksheet.max_column or 0
                if rows > self.limits.xlsx_max_rows_per_sheet:
                    raise StatementResourceLimitError(
                        "XLSX sheet exceeds the configured row extraction limit."
                    )
                if columns > self.limits.xlsx_max_columns_per_sheet:
                    raise StatementResourceLimitError(
                        "XLSX sheet exceeds the configured column extraction limit."
                    )
                declared_cells = rows * columns
                if total_cells + declared_cells > self.limits.xlsx_max_cells:
                    raise StatementResourceLimitError(
                        "XLSX exceeds the configured total cell extraction limit."
                    )
                table, actual_cells = worksheet_table(
                    worksheet.iter_rows(values_only=True),
                    max_rows=self.limits.xlsx_max_rows_per_sheet,
                    max_columns=self.limits.xlsx_max_columns_per_sheet,
                    max_cells=self.limits.xlsx_max_cells - total_cells,
                )
                total_cells += max(declared_cells, actual_cells)
                text_by_page.append(table_as_text(table))
                tables_by_page.append(
                    ExtractedStatementPageTables(
                        page_number=sheet_index,
                        tables=[table] if table else [],
                    )
                )

            return ExtractedStatement(
                text_by_page=text_by_page,
                tables_by_page=tables_by_page,
                metadata={
                    "extractor_name": self.parser_name,
                    "extractor_version": self.parser_version,
                    "source_format": "xlsx",
                    "sheet_names": list(workbook.sheetnames),
                },
            )
        finally:
            workbook.close()


def validate_xlsx_archive(file_path: Path, max_uncompressed_bytes: int) -> None:
    try:
        with ZipFile(file_path) as archive:
            uncompressed_bytes = 0
            for member in archive.infolist():
                uncompressed_bytes += member.file_size
                if uncompressed_bytes > max_uncompressed_bytes:
                    raise StatementResourceLimitError(
                        "XLSX exceeds the configured uncompressed size limit."
                    )
    except BadZipFile as error:
        raise InvalidFileException("XLSX is not a valid ZIP archive.") from error


def worksheet_table(
    rows: Any,
    *,
    max_rows: int,
    max_columns: int,
    max_cells: int,
) -> tuple[list[list[str | None]], int]:
    table: list[list[str | None]] = []
    cell_count = 0
    for row_index, row in enumerate(rows, start=1):
        if row_index > max_rows:
            raise StatementResourceLimitError(
                "XLSX sheet exceeds the configured row extraction limit."
            )
        if len(row) > max_columns:
            raise StatementResourceLimitError(
                "XLSX sheet exceeds the configured column extraction limit."
            )
        cell_count += len(row)
        if cell_count > max_cells:
            raise StatementResourceLimitError(
                "XLSX exceeds the configured total cell extraction limit."
            )
        normalized_row = [cell_as_text(value) for value in row]
        if any(value is not None for value in normalized_row):
            table.append(trim_trailing_empty_cells(normalized_row))
    return table, cell_count


def cell_as_text(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    text = str(value)
    return text if text.strip() else None


def trim_trailing_empty_cells(row: list[str | None]) -> list[str | None]:
    while row and row[-1] is None:
        row.pop()
    return row


def table_as_text(table: list[list[str | None]]) -> str:
    return "\n".join("\t".join(value or "" for value in row) for row in table)
