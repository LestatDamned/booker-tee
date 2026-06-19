from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.features.imports.infrastructure.extraction.extracted_statement import (
    ExtractedStatement,
    ExtractedStatementPageTables,
)


class OpenPyxlStatementExtractor:
    parser_name = "openpyxl_raw_extractor"
    parser_version = "0.1"

    def extract(self, file_path: Path) -> ExtractedStatement:
        workbook = load_workbook(
            file_path,
            read_only=True,
            data_only=True,
        )
        try:
            text_by_page: list[str] = []
            tables_by_page: list[ExtractedStatementPageTables] = []
            for sheet_index, worksheet in enumerate(workbook.worksheets, start=1):
                table = worksheet_table(worksheet.iter_rows(values_only=True))
                text_by_page.append(table_as_text(table))
                tables_by_page.append(
                    ExtractedStatementPageTables(
                        page_number=sheet_index,
                        tables=[table] if table else [],
                    )
                )

            return ExtractedStatement(
                text_by_page=text_by_page,
                tables_by_page=tables_by_page,
                metadata={
                    "extractor_name": self.parser_name,
                    "extractor_version": self.parser_version,
                    "source_format": "xlsx",
                    "sheet_names": list(workbook.sheetnames),
                },
            )
        finally:
            workbook.close()


def worksheet_table(rows: Any) -> list[list[str | None]]:
    table: list[list[str | None]] = []
    for row in rows:
        normalized_row = [cell_as_text(value) for value in row]
        if any(value is not None for value in normalized_row):
            table.append(trim_trailing_empty_cells(normalized_row))
    return table


def cell_as_text(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    text = str(value)
    return text if text.strip() else None


def trim_trailing_empty_cells(row: list[str | None]) -> list[str | None]:
    while row and row[-1] is None:
        row.pop()
    return row


def table_as_text(table: list[list[str | None]]) -> str:
    return "\n".join("\t".join(value or "" for value in row) for row in table)
